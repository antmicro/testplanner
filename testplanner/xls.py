import re
from typing import Tuple, Union

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet


class XLSX_writer:
    re_intent = re.compile(r"Intent:\n((^.+(\n|$))+)", flags=re.MULTILINE)
    re_stimulus = re.compile(r"Stimulus:\n((^.+(\n|$))+)", flags=re.MULTILINE)
    re_check = re.compile(r"Check:\n((^.+(\n|$))+)", flags=re.MULTILINE)
    re_blanklines = re.compile(r"\n\n(\n+)", flags=re.MULTILINE)

    def __init__(self, fp):
        self.fp = fp
        self.wb = load_workbook(fp)
        self.template_sheet = self.wb[self.wb.sheetnames[0]]
        calc_dim_y = re.sub(
            "[a-zA-Z]", "", self.template_sheet.calculate_dimension().split(":")[1]
        )
        self.tpl_entry_idx = int(calc_dim_y) + 1
        self.active_worksheet: Worksheet | None = None
        self.xls_column_map = {
            "name": "B",
            "type": "C",
            "metric": "D",
            "testbench": "E",
            "intent": "F",
            "stimulus_procedure": "G",
            "checking_mechanism": "H",
            "coverpoints": "I",
            "assignee": "J",
            "milestone": "K",
            "priority": "L",
            "status": "M",
            "effort_estimate": "N",
            "remaining": "O",
            "done": "P",
            "comments": "Q",
        }
        template_id_cell_pos = self.find_first_empty_cell()
        if template_id_cell_pos:
            self.template_id_pos_x = template_id_cell_pos[0]
            self.template_id_pos_y = template_id_cell_pos[1]

    def find_first_empty_cell(self) -> Union[Tuple[int, int], None]:
        for row in self.template_sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    return cell.row, cell.column
        return None

    def save(self):
        self.wb.save(self.fp)

    def create_or_select_sheet(self, name: str | None):
        """
        Creates a sheet from the template sheet and selects it as the active
        sheet. If a sheet with the specified name exists, it will be selected
        as active but not overwritten.
        """
        if name is None:
            raise RuntimeError("Cannot create an unnamed worksheet")
        if name in self.wb:
            self.active_worksheet = self.wb[name]
        else:
            self.active_worksheet = self.wb.copy_worksheet(self.template_sheet)
            self.active_worksheet.title = name
            self.active_worksheet.cell(
                self.template_id_pos_x, self.template_id_pos_y
            ).value = name

    def parse_standard_description(
        self, desc: str, headers: dict[str, str], clean_description: bool = False
    ) -> Tuple[dict[str, str], str]:
        """
        Processes a testplans description according to 'header' data

        Keyword arguments:
        desc -- the description string
        headers -- a dict describing the relation between
                   searched keywords and testplan columns
        clean_description -- whether to remove the found sections
                             from the description (default: False)

        """
        strs = {}
        clean_desc = desc
        parts = desc.split("\n\n")

        # Storing indexes to parts that have to be removed -
        # when it was done during the loop, it lead to skipping
        parts_to_remove = []
        state = None

        for part in parts:
            for header in headers.keys():
                if part.lstrip().startswith(header):
                    strs[headers[header]] = part.lstrip().removeprefix(header + ":\n")
                    state = header
                    if clean_description:
                        parts_to_remove.append(parts.index(part))
                    break
            else:
                if state:
                    strs[headers[state]] += "\n\n" + part
                    if clean_description:
                        parts_to_remove.append(parts.index(part))
        if clean_description:
            for ptr in sorted(parts_to_remove)[::-1]:
                parts.pop(ptr)
            clean_desc = "\n\n".join(parts)

        return strs, clean_desc

    def testplan_add_entry(
        self,
        name: str,
        data: dict[str, str],
        comment: str = "",
    ):
        """
        Adds an entry at 'self.tpl_entry_idx' containing passed data points
        """
        if self.active_worksheet is not None:
            self.active_worksheet.insert_rows(self.tpl_entry_idx)
            self.active_worksheet[
                self.xls_column_map["name"] + str(self.tpl_entry_idx)
            ] = name
            for key, value in data.items():
                self.active_worksheet[
                    self.xls_column_map[key] + str(self.tpl_entry_idx)
                ] = value
            self.active_worksheet[
                self.xls_column_map["comments"] + str(self.tpl_entry_idx)
            ] = comment
        else:
            raise RuntimeError("No worksheet was selected to be active")

    def testplan_append_to_entry_col(
        self, col: str, content: str, entry_key: str, entry_key_col: str = "name"
    ):
        """
        Finds a row by 'entry_key' in column 'entry_key_col' and within that row,
        puts 'content' str into the 'col' column
        """
        assert col in self.xls_column_map.keys()
        found = False
        for row in list(self.active_worksheet.rows)[self.tpl_entry_idx - 1 :]:
            if (
                row[
                    column_index_from_string(self.xls_column_map[entry_key_col]) - 1
                ].value
                == entry_key
            ):
                row[
                    column_index_from_string(self.xls_column_map[col]) - 1
                ].value = content
                found = True
        assert found, f"row for {entry_key} was not found!"

    def embolden_line(self, txt: str, lineno: int = 0):
        """
        Emboldens a selected line from a string and returns
        a rich-text variant of that string
        """
        txt_split = txt.split("\n")
        emboldened = CellRichText(
            "\n".join(txt_split[:lineno]),
            TextBlock(InlineFont(b=True), txt_split[lineno] + "\n"),
            "\n".join(txt_split[lineno + 1 :]),
        )
        return emboldened

    def format_string_columns(self):
        """
        Iterates through columns in need for formatting due to
        content length and calculates an acceptable width for them
        """
        columns_to_format = [
            self.xls_column_map["name"],
            self.xls_column_map["testbench"],
            self.xls_column_map["intent"],
            self.xls_column_map["comments"],
            self.xls_column_map["stimulus_procedure"],
            self.xls_column_map["checking_mechanism"],
            self.xls_column_map["status"],
            self.xls_column_map["comments"],
        ]
        columns_to_hide = [
            # self.xls_column_map["comments"],
        ]
        for column in self.active_worksheet.columns:
            if column[self.tpl_entry_idx - 1].column_letter in columns_to_format:
                max_length = 0
                for cell in column:
                    if cell.value:
                        lines = str(cell.value).split("\n")
                        for line in lines:
                            if len(line) > max_length:
                                max_length = len(line)
                adjusted_width = max_length + 2
                self.active_worksheet.column_dimensions[
                    column[self.tpl_entry_idx - 1].column_letter
                ].width = adjusted_width
            if column[self.tpl_entry_idx - 1].column_letter in columns_to_hide:
                self.active_worksheet.column_dimensions[
                    column[self.tpl_entry_idx - 1].column_letter
                ].hidden = True
